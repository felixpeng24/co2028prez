"""
Build the final Stompers stats workbook for Don't Die Week awards.

Reads:
  - Stompers Winners.xlsx     (Felix's manually-tracked ground truth, Sept 1 - Jan 31)
  - felix-friends-room-top3-2025-09-01-to-2026-04-23.xlsx  (app export)

Writes:
  - Stompers Final Stats.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from collections import defaultdict, Counter

# === Configuration ===

NON_MUDD_USERNAMES = {
    'penneyli', 'jessewu', 'reinan', 'siran', 'jamespeng', 'jasminetan',
    'edithchern', 'charlottep0', 'shannawu', 'mollyhuang2', 'danielxu7',
    'hankswang', 'andyli9', 'shirleyzha4', 'stevenzhan6',
}

NON_MUDD_DISPLAY = {
    'penneyli': 'Penney Li', 'jessewu': 'Jesse Wu', 'reinan': 'reina n',
    'siran': 'Sir An', 'jamespeng': 'James Peng', 'jasminetan': 'jasmine tan',
    'edithchern': 'edith chern', 'charlottep0': 'Charlotte Peng',
    'shannawu': 'Shanna Wu', 'mollyhuang2': 'Molly Huang', 'danielxu7': 'Daniel Xu',
    'hankswang': 'Hanks Wang', 'andyli9': 'Andy Li',
    'shirleyzha4': 'Shirley Zhang', 'stevenzhan6': 'Steven Zhang',
}

CC_USERNAMES = {'brannanr', 'jerryxu', 'arnavshetty', 'sophiesale', 'oliviaback'}

# Manual sheet display name -> app username
NAME_TO_USERNAME = {
    'Brannan Rosenfeld': 'brannanr', 'Leah Uriarte': 'leahtiger',
    'Felix Peng': 'felixp', 'Lexie Stinson': 'lexie1',
    'Andy Maldonado': 'andymaldon', 'Charlotte Almond': 'charlottea73',
    'Sophie Wang': 'sophiewang', 'Henry Allen': 'henryallen',
    'Silas Brock': 'silasbrock', 'Dylan Rainman': 'dylanrainm',
    'Arnav Shetty': 'arnavshetty', 'Jerry Xu': 'jerryxu',
    'Sophie Saleh': 'sophiesale', 'Olivia Backholm': 'oliviaback',
    'Natalie Ko': 'ko12nat', 'Megan Tran': 'megantran',
    'Penelope Rodriguez': 'penelopero0', 'Lucas Giraldo': 'lucasgiral',
    'Carolyn Davis': 'carolyndav0', 'Alana Jonick': 'alanajonic',
    'Romeo Velarde-Alvarez': 'natavelarde', 'Logan Mansfield': 'loganmansf',
    'Talia Green': 'taliagreen', 'Abby Sandorff': 'abbysandor',
    'Kate Risse': 'katerisse', 'Samara Baidwan': 'samarabaid',
    'Jack Van der Reis': 'jackvander1', 'Angie Hou': 'angiehou',
    'Alex B': 'alexb79',
}

# Canonical display name (preferred form) per username
CANONICAL_DISPLAY = {
    'henryallen': 'Henry Allen', 'charlottea73': 'Charlotte Almond',
    'arnavshetty': 'Arnav Shetty', 'jerryxu': 'Jerry Xu',
    'leahtiger': 'Leah Uriarte', 'silasbrock': 'Silas Brock',
    'andymaldon': 'Andy Maldonado', 'felixp': 'Felix Peng',
    'ko12nat': 'Natalie Ko', 'brannanr': 'Brannan Rosenfeld',
    'dylanrainm': 'Dylan Rainman', 'abbysandor': 'Abby Sandorff',
    'alanajonic': 'Alana Jonick', 'penelopero0': 'Penelope Rodriguez',
    'jackvander1': 'Jack Van der Reis', 'carolyndav0': 'Carolyn Davis',
    'samarabaid': 'Samara Baidwan', 'loganmansf': 'Logan Mansfield',
    'angiehou': 'Angie Hou', 'taliagreen': 'Talia Green',
    'lexie1': 'Lexie Stinson', 'lucasgiral': 'Lucas Giraldo',
    'sophiesale': 'Sophie Saleh', 'sophiewang': 'Sophie Wang',
    'megantran': 'Megan Tran', 'oliviaback': 'Olivia Backholm',
    'natavelarde': 'Romeo Velarde-Alvarez', 'katerisse': 'Kate Risse',
    'alexb79': 'Alex B', 'katiec4': 'Katie C',
}

MANUAL_FILE = 'Stompers Winners.xlsx'
APP_FILE = 'felix-friends-room-top3-2025-09-01-to-2026-04-23.xlsx'
OUTPUT_FILE = 'Stompers Final Stats.xlsx'

MANUAL_END = date(2026, 1, 31)  # last day Felix manually tracked

# Strings in manual sheet that aren't names (CC notes)
MANUAL_NOISE = {'wow', 'finally', 'wow x2', 'wow x3', 'wow x4', 'wow x5', 'wow x6',
                'wow x7', 'wow x8', 'wow x9', 'wow x10', 'wow x11', 'wow x12',
                'wow x13', 'wow x14', 'wow x15', 'wow x16', 'wow x17', 'wow x18',
                'wow x19'}


def is_noise(s):
    if s is None:
        return True
    return str(s).strip().lower() in MANUAL_NOISE


# === Read app data ===

app_wb = openpyxl.load_workbook(APP_FILE, data_only=True)

# Day -> list of (rank, name, username, steps) for that day's top 3 from app
app_daily = {}  # date -> [(rank, name, username, steps), ...]
ws = app_wb['Top 3 By Game']
rows = list(ws.iter_rows(values_only=True))
for row in rows[1:]:
    date_str = row[0]
    d = datetime.strptime(date_str, '%Y-%m-%d').date()
    entries = []
    for rank, base in [(1, 6), (2, 10), (3, 14)]:
        name = row[base]
        username = row[base + 1]
        steps = row[base + 3]
        if name and steps:
            entries.append((rank, name, username, steps))
    app_daily[d] = entries

# Build username -> step count by date for quick lookup
username_day_steps = defaultdict(dict)  # username -> {date: steps}
for d, entries in app_daily.items():
    for rank, name, username, steps in entries:
        username_day_steps[username][d] = steps


# === Read manual data ===

manual_wb = openpyxl.load_workbook(MANUAL_FILE, data_only=True)
ws = manual_wb['Sheet1']
manual_rows = list(ws.iter_rows(values_only=True))

# date -> {'reg': [name1, name2, name3], 'cc': [name1, name2, name3]}
manual_daily = {}
for row in manual_rows[2:]:
    d = row[0]
    if not isinstance(d, datetime):
        continue
    d = d.date()
    if d > MANUAL_END:
        continue
    reg = [row[1], row[2], row[3]]
    cc = [row[7], row[8], row[9]]
    reg = [(n.strip() if n else None) for n in reg]
    cc = [(n.strip() if n and not is_noise(n) else None) for n in cc]
    if any(reg) or any(cc):
        manual_daily[d] = {'reg': reg, 'cc': cc}


# === Build unified daily filtered top 3 ===
# For Sept 1 - Jan 31: use manual data as ground truth (already CC-split, non-Mudd-omitted)
# For Feb 1 - Apr 23: use app data, filter non-Mudd, split CC

daily_filtered = {}  # date -> {'reg': [(name, steps_or_None)], 'cc': [(name, steps_or_None)]}

# Sept 1 - Jan 31 from manual
for d, day in manual_daily.items():
    reg_entries = []
    cc_entries = []
    for name in day['reg']:
        if name is None:
            reg_entries.append(None)
            continue
        username = NAME_TO_USERNAME.get(name)
        steps = None
        if username and d in username_day_steps.get(username, {}):
            steps = username_day_steps[username][d]
        reg_entries.append((name, username, steps))
    for name in day['cc']:
        if name is None:
            cc_entries.append(None)
            continue
        username = NAME_TO_USERNAME.get(name)
        steps = None
        if username and d in username_day_steps.get(username, {}):
            steps = username_day_steps[username][d]
        cc_entries.append((name, username, steps))
    daily_filtered[d] = {'reg': reg_entries, 'cc': cc_entries}

# Feb 1 - Apr 23 from app
for d, entries in app_daily.items():
    if d <= MANUAL_END:
        continue
    reg_entries = []
    cc_entries = []
    for rank, name, username, steps in entries:
        if username in NON_MUDD_USERNAMES:
            continue
        display = CANONICAL_DISPLAY.get(username, name)
        if username in CC_USERNAMES:
            cc_entries.append((display, username, steps))
        else:
            reg_entries.append((display, username, steps))
    # Pad to 3 with Nones (so columns line up)
    while len(reg_entries) < 3:
        reg_entries.append(None)
    while len(cc_entries) < 3:
        cc_entries.append(None)
    # Trim to 3 (rare: more than 3 Mudd-non-CC in app top 3, can't happen since app top 3 is 3)
    daily_filtered[d] = {'reg': reg_entries[:3], 'cc': cc_entries[:3]}


# === Compute points and stats ===

POINTS = {0: 3, 1: 2, 2: 1}  # rank index -> points

# Per-username stats
reg_stats = defaultdict(lambda: {'first': 0, 'second': 0, 'third': 0, 'days': set(), 'best_steps': 0, 'best_steps_date': None})
cc_stats = defaultdict(lambda: {'first': 0, 'second': 0, 'third': 0, 'days': set(), 'best_steps': 0, 'best_steps_date': None})

for d, day in daily_filtered.items():
    for idx, entry in enumerate(day['reg']):
        if entry is None:
            continue
        name, username, steps = entry
        if username is None:
            # Unknown person from manual sheet, store under display name
            key = ('UNKNOWN', name)
        else:
            key = username
        s = reg_stats[key]
        if idx == 0:
            s['first'] += 1
        elif idx == 1:
            s['second'] += 1
        elif idx == 2:
            s['third'] += 1
        s['days'].add(d)
        if steps and steps > s['best_steps']:
            s['best_steps'] = steps
            s['best_steps_date'] = d
    for idx, entry in enumerate(day['cc']):
        if entry is None:
            continue
        name, username, steps = entry
        if username is None:
            key = ('UNKNOWN', name)
        else:
            key = username
        s = cc_stats[key]
        if idx == 0:
            s['first'] += 1
        elif idx == 1:
            s['second'] += 1
        elif idx == 2:
            s['third'] += 1
        s['days'].add(d)
        if steps and steps > s['best_steps']:
            s['best_steps'] = steps
            s['best_steps_date'] = d


def display_name(key):
    if isinstance(key, tuple):
        return key[1]
    return CANONICAL_DISPLAY.get(key, key)


def points(stats):
    return stats['first'] * 3 + stats['second'] * 2 + stats['third'] * 1


# === Compute fun awards ===

# Single-Day King: highest single-day step count (Mudd, any pool)
all_mudd_day_steps = []  # (steps, date, name)
for username, days in username_day_steps.items():
    if username in NON_MUDD_USERNAMES:
        continue
    for d, steps in days.items():
        all_mudd_day_steps.append((steps, d, CANONICAL_DISPLAY.get(username, username)))
all_mudd_day_steps.sort(reverse=True)

# 30K Club
thirty_k_club = [(s, d, n) for (s, d, n) in all_mudd_day_steps if s >= 30000]

# Early Bird (most podiums in September) - regular pool
sept_counts = Counter()
for d, day in daily_filtered.items():
    if d.month != 9 or d.year != 2025:
        continue
    for entry in day['reg']:
        if entry is None:
            continue
        _, username, _ = entry
        if username:
            sept_counts[username] += 1

# Spring Surge (most podiums Feb 1+)
spring_counts = Counter()
for d, day in daily_filtered.items():
    if d < date(2026, 2, 1):
        continue
    for entry in day['reg']:
        if entry is None:
            continue
        _, username, _ = entry
        if username:
            spring_counts[username] += 1

# Hot Streak: longest consecutive-day podium streak (regular pool)
def longest_streak_per_user(pool='reg'):
    days_by_user = defaultdict(set)
    for d, day in daily_filtered.items():
        for entry in day[pool]:
            if entry is None:
                continue
            _, username, _ = entry
            if username:
                days_by_user[username].add(d)
    streaks = {}
    for username, days in days_by_user.items():
        sorted_days = sorted(days)
        best = 1
        cur = 1
        for i in range(1, len(sorted_days)):
            if (sorted_days[i] - sorted_days[i - 1]).days == 1:
                cur += 1
                best = max(best, cur)
            else:
                cur = 1
        streaks[username] = best
    return streaks

reg_streaks = longest_streak_per_user('reg')

# Comeback / Ghost: longest gap between two consecutive podiums (min 3 podiums total)
def longest_gap_per_user(pool='reg'):
    days_by_user = defaultdict(list)
    for d, day in daily_filtered.items():
        for entry in day[pool]:
            if entry is None:
                continue
            _, username, _ = entry
            if username:
                days_by_user[username].append(d)
    gaps = {}
    for username, days in days_by_user.items():
        if len(days) < 3:
            continue
        sorted_days = sorted(days)
        max_gap = 0
        max_gap_pair = None
        for i in range(1, len(sorted_days)):
            gap = (sorted_days[i] - sorted_days[i - 1]).days
            if gap > max_gap:
                max_gap = gap
                max_gap_pair = (sorted_days[i - 1], sorted_days[i])
        gaps[username] = (max_gap, max_gap_pair)
    return gaps

reg_gaps = longest_gap_per_user('reg')

# Dynamic Duo: pair that podiumed together most days (regular pool)
duo_counts = Counter()
for d, day in daily_filtered.items():
    on_podium = []
    for entry in day['reg']:
        if entry is None:
            continue
        _, username, _ = entry
        if username:
            on_podium.append(username)
    on_podium = list(set(on_podium))
    on_podium.sort()
    for i in range(len(on_podium)):
        for j in range(i + 1, len(on_podium)):
            duo_counts[(on_podium[i], on_podium[j])] += 1

# The Sniper: smallest-margin 1st place wins
# Need to compute margin between 1st and 2nd. Can only do this from app data
# (manual data doesn't have steps). So sniper is computed from app data filtered.
sniper_data = []  # (margin, date, winner_name, runnerup_name)
for d, entries in app_daily.items():
    # Filter to Mudd, keep order
    filtered = [(rank, name, username, steps) for (rank, name, username, steps) in entries
                if username not in NON_MUDD_USERNAMES]
    if len(filtered) < 2:
        continue
    # First Mudd vs second Mudd
    first = filtered[0]
    second = filtered[1]
    margin = first[3] - second[3]
    sniper_data.append((margin, d, CANONICAL_DISPLAY.get(first[2], first[1]),
                        CANONICAL_DISPLAY.get(second[2], second[1]), first[3], second[3]))
sniper_data.sort()

# Late Bloomer: latest first-podium date for someone with >=2 podiums total
first_podium = {}
podium_count_total = Counter()
for d in sorted(daily_filtered.keys()):
    day = daily_filtered[d]
    for entry in day['reg'] + day['cc']:
        if entry is None:
            continue
        _, username, _ = entry
        if username:
            podium_count_total[username] += 1
            if username not in first_podium:
                first_podium[username] = d

late_bloomer_candidates = [(d, u) for u, d in first_podium.items()
                            if podium_count_total[u] >= 2]
late_bloomer_candidates.sort(reverse=True)

# Most Improved: biggest jump in monthly podium count (regular pool)
monthly_counts = defaultdict(Counter)  # username -> Counter of (year, month) -> count
for d, day in daily_filtered.items():
    for entry in day['reg']:
        if entry is None:
            continue
        _, username, _ = entry
        if username:
            monthly_counts[username][(d.year, d.month)] += 1

# For each user, find biggest month-over-month increase
most_improved = []  # (jump, username, from_month, to_month, from_count, to_count)
all_months = sorted({(d.year, d.month) for d in daily_filtered.keys()})
for username, counts in monthly_counts.items():
    if sum(counts.values()) < 5:
        continue  # Need a baseline
    best_jump = 0
    best_pair = None
    for i in range(1, len(all_months)):
        prev_m = all_months[i - 1]
        cur_m = all_months[i]
        jump = counts[cur_m] - counts[prev_m]
        if jump > best_jump:
            best_jump = jump
            best_pair = (prev_m, cur_m, counts[prev_m], counts[cur_m])
    if best_pair:
        most_improved.append((best_jump, username, *best_pair))
most_improved.sort(reverse=True)


# === Build the output workbook ===

out = openpyxl.Workbook()
out.remove(out.active)  # delete default sheet

HEADER_FILL = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, size=14, color="D94F4F")
TITLE_FONT = Font(bold=True, size=16, color="28587B")
SUBTITLE_FONT = Font(italic=True, size=10, color="666666")


def style_header_row(ws, row_idx, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
            try:
                v = str(cell.value) if cell.value is not None else ''
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


# ---- Sheet 1: Awards ----
ws = out.create_sheet('Awards')
r = 1
ws.cell(row=r, column=1, value='STOMPERS YEAR-END AWARDS').font = TITLE_FONT
r += 1
ws.cell(row=r, column=1, value='Class of 2028 | September 1, 2025 – April 23, 2026').font = SUBTITLE_FONT
r += 2

ws.cell(row=r, column=1, value='HEADLINE AWARDS (gift winners)').font = SECTION_FONT
r += 2

# Stompers Champion (top 3 by points, regular pool)
ws.cell(row=r, column=1, value='STOMPERS CHAMPION (by points: 3 / 2 / 1)').font = Font(bold=True, size=12)
r += 1
ws.cell(row=r, column=1, value='Place')
ws.cell(row=r, column=2, value='Name')
ws.cell(row=r, column=3, value='Points')
ws.cell(row=r, column=4, value='1st Places')
ws.cell(row=r, column=5, value='2nd Places')
ws.cell(row=r, column=6, value='3rd Places')
style_header_row(ws, r, 6)
r += 1

ranked_reg = sorted(reg_stats.items(), key=lambda kv: (-points(kv[1]), -kv[1]['first'], -kv[1]['second']))
medals = {0: 'Gold', 1: 'Silver', 2: 'Bronze'}
for i, (key, s) in enumerate(ranked_reg[:3]):
    ws.cell(row=r, column=1, value=medals[i])
    ws.cell(row=r, column=2, value=display_name(key))
    ws.cell(row=r, column=3, value=points(s))
    ws.cell(row=r, column=4, value=s['first'])
    ws.cell(row=r, column=5, value=s['second'])
    ws.cell(row=r, column=6, value=s['third'])
    r += 1
r += 1

# Cross Country MVP
ws.cell(row=r, column=1, value='CROSS COUNTRY MVP').font = Font(bold=True, size=12)
r += 1
ws.cell(row=r, column=1, value='Place')
ws.cell(row=r, column=2, value='Name')
ws.cell(row=r, column=3, value='Points')
ws.cell(row=r, column=4, value='1st Places')
ws.cell(row=r, column=5, value='2nd Places')
ws.cell(row=r, column=6, value='3rd Places')
style_header_row(ws, r, 6)
r += 1

ranked_cc = sorted(cc_stats.items(), key=lambda kv: (-points(kv[1]), -kv[1]['first']))
for i, (key, s) in enumerate(ranked_cc[:3]):
    ws.cell(row=r, column=1, value=medals[i])
    ws.cell(row=r, column=2, value=display_name(key))
    ws.cell(row=r, column=3, value=points(s))
    ws.cell(row=r, column=4, value=s['first'])
    ws.cell(row=r, column=5, value=s['second'])
    ws.cell(row=r, column=6, value=s['third'])
    r += 1
r += 2

ws.cell(row=r, column=1, value='FUN / WEIRD AWARDS (no gift, all glory)').font = SECTION_FONT
r += 2

def add_simple_award(title, winner_label, winner_value, detail=''):
    global r
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=11, color="E87D2F")
    r += 1
    ws.cell(row=r, column=1, value=winner_label).font = Font(bold=True)
    ws.cell(row=r, column=2, value=winner_value)
    if detail:
        ws.cell(row=r, column=3, value=detail)
    r += 2

# Bronze King/Queen
ranked_bronze = sorted(reg_stats.items(), key=lambda kv: -kv[1]['third'])
if ranked_bronze:
    k, s = ranked_bronze[0]
    add_simple_award('BRONZE KING/QUEEN (most 3rd-place finishes)',
                     display_name(k), f"{s['third']} bronze finishes",
                     f"({s['first']} firsts, {s['second']} seconds for context)")

# Single-Day King
if all_mudd_day_steps:
    s, d, n = all_mudd_day_steps[0]
    add_simple_award('SINGLE-DAY KING (highest one-day step count)',
                     n, f"{s:,} steps", f"on {d.strftime('%B %d, %Y')}")

# 30K Club
if thirty_k_club:
    ws.cell(row=r, column=1, value='30K CLUB (broke 30,000 steps in a single day)').font = Font(bold=True, size=11, color="E87D2F")
    r += 1
    ws.cell(row=r, column=1, value='Name')
    ws.cell(row=r, column=2, value='Steps')
    ws.cell(row=r, column=3, value='Date')
    style_header_row(ws, r, 3)
    r += 1
    seen = set()
    for s, d, n in thirty_k_club:
        # Multiple entries possible; show all for full glory
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=s)
        ws.cell(row=r, column=3, value=d.strftime('%Y-%m-%d'))
        r += 1
    r += 1

# Early Bird
if sept_counts:
    top = sept_counts.most_common(1)[0]
    add_simple_award('EARLY BIRD (most podiums in September 2025)',
                     CANONICAL_DISPLAY.get(top[0], top[0]), f"{top[1]} podiums in September")

# Spring Surge
if spring_counts:
    top = spring_counts.most_common(1)[0]
    add_simple_award('SPRING SURGE (most podiums Feb 1 onwards)',
                     CANONICAL_DISPLAY.get(top[0], top[0]), f"{top[1]} podiums in Feb–April")

# Hot Streak
if reg_streaks:
    top_user = max(reg_streaks.items(), key=lambda kv: kv[1])
    add_simple_award('HOT STREAK (longest consecutive-day podium streak)',
                     CANONICAL_DISPLAY.get(top_user[0], top_user[0]),
                     f"{top_user[1]} days in a row")

# Comeback Award (longest gap)
if reg_gaps:
    top_gap = max(reg_gaps.items(), key=lambda kv: kv[1][0])
    username, (gap, pair) = top_gap
    add_simple_award('COMEBACK AWARD (longest dry spell between podiums, 3+ podiums total)',
                     CANONICAL_DISPLAY.get(username, username),
                     f"{gap}-day gap",
                     f"between {pair[0].strftime('%b %d')} and {pair[1].strftime('%b %d')}")

# Dynamic Duo
if duo_counts:
    top_duo = duo_counts.most_common(1)[0]
    (u1, u2), c = top_duo
    add_simple_award('DYNAMIC DUO (pair that podiumed together most days)',
                     f"{CANONICAL_DISPLAY.get(u1, u1)} & {CANONICAL_DISPLAY.get(u2, u2)}",
                     f"{c} days on the podium together")

# The Sniper (smallest 1st-place margin among Mudd-only filtered)
if sniper_data:
    margin, d, winner, runnerup, w_steps, r_steps = sniper_data[0]
    add_simple_award('THE SNIPER (smallest 1st-place margin)',
                     winner, f"won by just {margin} steps",
                     f"({w_steps:,} vs {r_steps:,} for {runnerup} on {d.strftime('%b %d, %Y')})")

# Late Bloomer
if late_bloomer_candidates:
    d, username = late_bloomer_candidates[0]
    add_simple_award('LATE BLOOMER (latest first-podium of someone with multiple podiums)',
                     CANONICAL_DISPLAY.get(username, username),
                     f"first podium on {d.strftime('%B %d, %Y')}",
                     f"({podium_count_total[username]} total podiums)")

# Most Improved
if most_improved:
    jump, username, prev_m, cur_m, prev_c, cur_c = most_improved[0]
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    add_simple_award('MOST IMPROVED (biggest month-over-month podium jump)',
                     CANONICAL_DISPLAY.get(username, username),
                     f"+{jump} podiums",
                     f"({prev_c} in {months[prev_m[1]-1]} → {cur_c} in {months[cur_m[1]-1]})")

# Don't Die Award
ws.cell(row=r, column=1, value="DON'T DIE AWARD (survived sophomore year)").font = Font(bold=True, size=11, color="E87D2F")
r += 1
ws.cell(row=r, column=1, value='Winner').font = Font(bold=True)
ws.cell(row=r, column=2, value='Everyone who shows up to Don\'t Die Week. You did it.')
r += 2

autosize(ws)


# ---- Sheet 2: Points Leaderboard ----
ws = out.create_sheet('Points Leaderboard')
ws.cell(row=1, column=1, value='REGULAR POOL — POINTS LEADERBOARD').font = TITLE_FONT
ws.cell(row=2, column=1, value='Points = (1st × 3) + (2nd × 2) + (3rd × 1). Cross country runners scored separately.').font = SUBTITLE_FONT
ws.cell(row=4, column=1, value='Rank')
ws.cell(row=4, column=2, value='Name')
ws.cell(row=4, column=3, value='Points')
ws.cell(row=4, column=4, value='1st')
ws.cell(row=4, column=5, value='2nd')
ws.cell(row=4, column=6, value='3rd')
ws.cell(row=4, column=7, value='Total Podiums')
ws.cell(row=4, column=8, value='Distinct Days')
ws.cell(row=4, column=9, value='Best Single-Day Steps')
ws.cell(row=4, column=10, value='Best Day')
style_header_row(ws, 4, 10)

r = 5
for i, (key, s) in enumerate(ranked_reg, start=1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=display_name(key))
    ws.cell(row=r, column=3, value=points(s))
    ws.cell(row=r, column=4, value=s['first'])
    ws.cell(row=r, column=5, value=s['second'])
    ws.cell(row=r, column=6, value=s['third'])
    ws.cell(row=r, column=7, value=s['first'] + s['second'] + s['third'])
    ws.cell(row=r, column=8, value=len(s['days']))
    ws.cell(row=r, column=9, value=s['best_steps'] if s['best_steps'] else '')
    ws.cell(row=r, column=10, value=s['best_steps_date'].strftime('%Y-%m-%d') if s['best_steps_date'] else '')
    r += 1

autosize(ws)


# ---- Sheet 3: Cross Country Leaderboard ----
ws = out.create_sheet('Cross Country Leaderboard')
ws.cell(row=1, column=1, value='CROSS COUNTRY POOL — POINTS LEADERBOARD').font = TITLE_FONT
ws.cell(row=2, column=1, value='CC runners: Brannan Rosenfeld, Jerry Xu, Arnav Shetty, Sophie Saleh, Olivia Backholm').font = SUBTITLE_FONT
ws.cell(row=4, column=1, value='Rank')
ws.cell(row=4, column=2, value='Name')
ws.cell(row=4, column=3, value='Points')
ws.cell(row=4, column=4, value='1st')
ws.cell(row=4, column=5, value='2nd')
ws.cell(row=4, column=6, value='3rd')
ws.cell(row=4, column=7, value='Total Podiums')
ws.cell(row=4, column=8, value='Distinct Days')
ws.cell(row=4, column=9, value='Best Single-Day Steps')
ws.cell(row=4, column=10, value='Best Day')
style_header_row(ws, 4, 10)

r = 5
for i, (key, s) in enumerate(ranked_cc, start=1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=display_name(key))
    ws.cell(row=r, column=3, value=points(s))
    ws.cell(row=r, column=4, value=s['first'])
    ws.cell(row=r, column=5, value=s['second'])
    ws.cell(row=r, column=6, value=s['third'])
    ws.cell(row=r, column=7, value=s['first'] + s['second'] + s['third'])
    ws.cell(row=r, column=8, value=len(s['days']))
    ws.cell(row=r, column=9, value=s['best_steps'] if s['best_steps'] else '')
    ws.cell(row=r, column=10, value=s['best_steps_date'].strftime('%Y-%m-%d') if s['best_steps_date'] else '')
    r += 1

autosize(ws)


# ---- Sheet 4: Daily Top 3 (Filtered) ----
ws = out.create_sheet('Daily Top 3 (Filtered)')
ws.cell(row=1, column=1, value='DAILY TOP 3 — MUDD STUDENTS ONLY').font = TITLE_FONT
ws.cell(row=2, column=1, value='Sept 1 – Jan 31 from manual ground truth. Feb 1 – Apr 23 from app data with non-Mudd filtered.').font = SUBTITLE_FONT
ws.cell(row=4, column=1, value='Date')
ws.cell(row=4, column=2, value='Reg 1st')
ws.cell(row=4, column=3, value='Reg 2nd')
ws.cell(row=4, column=4, value='Reg 3rd')
ws.cell(row=4, column=5, value='CC 1st')
ws.cell(row=4, column=6, value='CC 2nd')
ws.cell(row=4, column=7, value='CC 3rd')
ws.cell(row=4, column=8, value='Source')
style_header_row(ws, 4, 8)

r = 5
for d in sorted(daily_filtered.keys()):
    day = daily_filtered[d]
    ws.cell(row=r, column=1, value=d.strftime('%Y-%m-%d'))
    for i, entry in enumerate(day['reg']):
        ws.cell(row=r, column=2 + i, value=entry[0] if entry else '')
    for i, entry in enumerate(day['cc']):
        ws.cell(row=r, column=5 + i, value=entry[0] if entry else '')
    ws.cell(row=r, column=8, value='manual' if d <= MANUAL_END else 'app-filtered')
    r += 1

autosize(ws)


# ---- Sheet 5: Single-Day Records ----
ws = out.create_sheet('Single-Day Records')
ws.cell(row=1, column=1, value='TOP 25 SINGLE-DAY STEP COUNTS (Mudd students)').font = TITLE_FONT
ws.cell(row=2, column=1, value='From app data; only days where the person was top 3 are visible.').font = SUBTITLE_FONT
ws.cell(row=4, column=1, value='Rank')
ws.cell(row=4, column=2, value='Name')
ws.cell(row=4, column=3, value='Steps')
ws.cell(row=4, column=4, value='Date')
ws.cell(row=4, column=5, value='Pool')
style_header_row(ws, 4, 5)

# Dedupe: for each day per person, take their max steps from app data
# (A person can only be in top 3 once per day)
unique_records = []
seen_per_user = defaultdict(set)
for s, d, n in all_mudd_day_steps:
    if (n, d) in seen_per_user[n]:
        continue
    seen_per_user[n].add((n, d))
    # Find their username to determine pool
    pool = 'Reg'
    for username, dn in CANONICAL_DISPLAY.items():
        if dn == n:
            if username in CC_USERNAMES:
                pool = 'CC'
            break
    unique_records.append((s, d, n, pool))
unique_records.sort(reverse=True)

r = 5
for i, (s, d, n, pool) in enumerate(unique_records[:25], start=1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=n)
    ws.cell(row=r, column=3, value=s)
    ws.cell(row=r, column=4, value=d.strftime('%Y-%m-%d'))
    ws.cell(row=r, column=5, value=pool)
    r += 1

autosize(ws)


# ---- Sheet 6: Monthly Breakdown ----
ws = out.create_sheet('Monthly Breakdown')
ws.cell(row=1, column=1, value='PODIUMS PER MONTH (Regular pool, Mudd only)').font = TITLE_FONT

months_in_range = sorted({(d.year, d.month) for d in daily_filtered.keys()})
month_labels = []
for y, m in months_in_range:
    label = f"{['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][m-1]} '{str(y)[-2:]}"
    month_labels.append(label)

ws.cell(row=3, column=1, value='Name')
for i, lbl in enumerate(month_labels):
    ws.cell(row=3, column=2 + i, value=lbl)
ws.cell(row=3, column=2 + len(month_labels), value='Total')
style_header_row(ws, 3, 2 + len(month_labels))

# Build per-user per-month counts (regular)
user_month = defaultdict(Counter)
for d, day in daily_filtered.items():
    for entry in day['reg']:
        if entry is None:
            continue
        _, username, _ = entry
        if username:
            user_month[username][(d.year, d.month)] += 1

# Sort users by total
sorted_users = sorted(user_month.keys(), key=lambda u: -sum(user_month[u].values()))

r = 4
for username in sorted_users:
    ws.cell(row=r, column=1, value=CANONICAL_DISPLAY.get(username, username))
    total = 0
    for i, ym in enumerate(months_in_range):
        c = user_month[username][ym]
        ws.cell(row=r, column=2 + i, value=c if c else '')
        total += c
    ws.cell(row=r, column=2 + len(month_labels), value=total)
    r += 1

autosize(ws)


# ---- Sheet 7: Excluded Names ----
ws = out.create_sheet('Excluded Names')
ws.cell(row=1, column=1, value='NAMES EXCLUDED FROM AWARDS (not Mudd students)').font = TITLE_FONT
ws.cell(row=2, column=1, value='Confirmed by Felix as not in the Class of 2028.').font = SUBTITLE_FONT
ws.cell(row=4, column=1, value='Display Name')
ws.cell(row=4, column=2, value='Username')
ws.cell(row=4, column=3, value='App Podiums (1st)')
ws.cell(row=4, column=4, value='App Podiums (2nd)')
ws.cell(row=4, column=5, value='App Podiums (3rd)')
style_header_row(ws, 4, 5)

# Compute their app podium counts from raw app data
excluded_counts = defaultdict(lambda: {'first': 0, 'second': 0, 'third': 0})
for d, entries in app_daily.items():
    for rank, name, username, steps in entries:
        if username in NON_MUDD_USERNAMES:
            if rank == 1:
                excluded_counts[username]['first'] += 1
            elif rank == 2:
                excluded_counts[username]['second'] += 1
            elif rank == 3:
                excluded_counts[username]['third'] += 1

r = 5
for username in sorted(NON_MUDD_USERNAMES):
    c = excluded_counts.get(username, {'first': 0, 'second': 0, 'third': 0})
    ws.cell(row=r, column=1, value=NON_MUDD_DISPLAY.get(username, username))
    ws.cell(row=r, column=2, value=username)
    ws.cell(row=r, column=3, value=c['first'])
    ws.cell(row=r, column=4, value=c['second'])
    ws.cell(row=r, column=5, value=c['third'])
    r += 1

autosize(ws)


# Save
out.save(OUTPUT_FILE)
print(f"Wrote {OUTPUT_FILE}")
print(f"  Regular pool people: {len(reg_stats)}")
print(f"  CC pool people: {len(cc_stats)}")
print(f"  Total days with data: {len(daily_filtered)}")
print(f"  30K club members (by entry): {len(thirty_k_club)}")
print()
print("Top 3 Regular by points:")
for i, (key, s) in enumerate(ranked_reg[:5], start=1):
    print(f"  {i}. {display_name(key)}: {points(s)} pts ({s['first']}/{s['second']}/{s['third']})")
print()
print("Top 3 CC by points:")
for i, (key, s) in enumerate(ranked_cc[:5], start=1):
    print(f"  {i}. {display_name(key)}: {points(s)} pts ({s['first']}/{s['second']}/{s['third']})")
